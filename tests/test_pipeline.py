import io
import sys
import unittest
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from export_utils import create_excel_report, create_research_package
from ml_pipeline import FEATURE_LABELS, clean_stock_data, evaluate_model, prepare_data, train_model


def market_data(start_year, end_year):
    dates = pd.bdate_range(f"{start_year}-01-01", f"{end_year}-12-31")
    rng = np.random.default_rng(42 + start_year + end_year)
    close = 100 * np.exp(np.cumsum(rng.normal(0.0002, 0.012, len(dates))))
    return pd.DataFrame({
        "DATE": dates,
        "Open Price": close * (1 + rng.normal(0, .002, len(dates))),
        "High Price": close * 1.01,
        "Low Price": close * .99,
        "Closing Price": close,
        "Total Traded Quantity": rng.integers(100_000, 2_000_000, len(dates)),
    })


class ResearchPipelineTests(unittest.TestCase):
    features = list(FEATURE_LABELS)

    def check_combined(self, start, end):
        prepared = prepare_data(market_data(start, end), self.features)
        self.assertEqual(prepared.train_data.Date.dt.year.max(), end - 1)
        self.assertEqual(prepared.test_data.Date.dt.year.unique().tolist(), [end])
        return prepared

    def test_a_2020_2024(self): self.check_combined(2020, 2024)
    def test_b_2020_2025(self): self.check_combined(2020, 2025)
    def test_c_2019_2024(self): self.check_combined(2019, 2024)

    def test_d_separate_file_context(self):
        train, test = market_data(2020, 2024), market_data(2025, 2025)
        prepared = prepare_data(train, self.features, test)
        first_test = prepared.test_data.iloc[0]
        expected = pd.concat([train, test]).sort_values("DATE")["Closing Price"].rolling(20).mean()
        expected.index = pd.concat([train, test]).sort_values("DATE")["DATE"].to_numpy()
        self.assertAlmostEqual(first_test.MA_20, expected.loc[first_test.Date], places=10)
        self.assertEqual(prepared.training_period, "2020–2024")
        self.assertEqual(prepared.test_period, "2025")

    def test_e_final_row_unlabelled(self):
        prepared = self.check_combined(2020, 2024)
        final = prepared.engineered_data.iloc[-1]
        self.assertTrue(pd.isna(final.Target))
        self.assertEqual(len(prepared.latest_unverified), 1)
        self.assertLess(prepared.test_evaluation.Date.max(), prepared.test_data.Date.max())

    def test_f_fit_allow_list_and_exports(self):
        prepared = self.check_combined(2020, 2024)
        chosen = ["MA_5", "RSI_14", "Momentum_5"]
        prepared = prepare_data(market_data(2020, 2024), chosen)
        model, params = train_model(prepared)
        self.assertEqual(list(model.feature_names_in_), chosen)
        forbidden = {"Next_Close", "Target", "Next_Day_Return"}
        self.assertFalse(forbidden.intersection(model.feature_names_in_))
        results = evaluate_model(model, prepared, .05)
        self.assertNotIn(prepared.test_data.Date.max(), set(results["predictions"].Date))
        excel = create_excel_report(prepared, results, params, .05)
        workbook = load_workbook(io.BytesIO(excel), read_only=False)
        self.assertEqual(workbook.sheetnames, ["Summary", "Raw Data", "Engineered Dataset", "Test Predictions", "Evaluation Metrics", "Benchmark Comparison", "Feature Importance", "Trading Performance", "Methodology"])
        package = create_research_package(prepared, results, model, params, .05)
        with zipfile.ZipFile(io.BytesIO(package)) as archive:
            names = set(archive.namelist())
            self.assertIn("research_package/random_forest_model.joblib", names)
            self.assertIn("research_package/charts/strategy_vs_buy_hold.png", names)

    def test_g_excel_imports_all_sheets(self):
        workbook = io.BytesIO()
        workbook.name = "history.xlsx"
        data = market_data(2020, 2024)
        chunks = np.array_split(data, 5)
        with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
            for index, chunk in enumerate(chunks, start=1):
                chunk.to_excel(writer, sheet_name=f"Sheet{index}", index=False)
        workbook.seek(0)
        cleaned = clean_stock_data(workbook)
        self.assertEqual(len(cleaned), len(data))
        self.assertEqual(cleaned.Date.min(), pd.Timestamp("2020-01-01"))
        self.assertEqual(cleaned.Date.max(), pd.Timestamp("2024-12-31"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
