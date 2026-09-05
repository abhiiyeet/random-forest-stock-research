"""In-memory research exports: CSV, XLSX, joblib model and ZIP package."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import json
import zipfile

import joblib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ml_pipeline import FEATURE_LABELS, PreparedData


NAVY = "24445C"
BLUE = "3478A8"
PALE_BLUE = "EAF2F7"
PALE_GREEN = "E8F3EC"
PALE_RED = "F8EAEA"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="D8DEE3")


def dataframe_csv(frame: pd.DataFrame) -> bytes:
    return frame.to_csv(index=False, date_format="%Y-%m-%d").encode("utf-8-sig")


def metrics_frame(results: dict) -> pd.DataFrame:
    return pd.DataFrame(
        {"Metric": list(results["metrics"]), "Value": list(results["metrics"].values())}
    )


def model_bytes(model) -> bytes:
    buffer = BytesIO()
    joblib.dump(model, buffer)
    return buffer.getvalue()


def _configuration(prepared: PreparedData, parameters: dict) -> dict:
    return {
        "model_name": "RandomForestClassifier",
        "selected_features": prepared.selected_features,
        "selected_feature_labels": [FEATURE_LABELS[x] for x in prepared.selected_features],
        "model_parameters": parameters,
        "training_date_range": prepared.training_period,
        "test_date_range": prepared.test_period,
        "training_observations": len(prepared.train_data),
        "test_observations": len(prepared.test_evaluation),
        "random_state": parameters.get("random_state", 42),
        "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
    }


def methodology_text(prepared: PreparedData, parameters: dict, transaction_cost: float) -> str:
    features = ", ".join(FEATURE_LABELS[f] for f in prepared.selected_features)
    return f"""AI in Algorithmic Trading: Next-Day Stock Direction Prediction Using Random Forest Classification

Research objective
Predict whether the next trading day's closing price is higher than today's closing price.

Input data
Daily Date, Open, High, Low, Close and Volume observations. Data are cleaned, de-duplicated by date and sorted chronologically.

Selected features
{features}

Feature formulas
MA_5 = rolling mean of Close over the current and previous 4 observations.
MA_20 = rolling mean of Close over the current and previous 19 observations.
RSI_14 = 100 - 100/(1 + 14-period average gain / 14-period average loss).
Momentum_5 = Close / Close.shift(5) - 1.
Volume_Change = Volume / Volume.shift(1) - 1; zero denominators are treated as missing.

Target
UP/1 when Next_Close > Close; DOWN/0 when Next_Close <= Close. A row without a genuine next close has no target and is excluded from evaluation.

Chronological split
Training period: {prepared.training_period}. Test period: {prepared.test_period}. No random split or shuffling is used. With one file, the latest calendar year is held out. With two files, the first is training and the later file is testing. Historical training prices provide rolling-feature context at the beginning of the test period, but test labels are never used in training.

Model
RandomForestClassifier with parameters: {json.dumps(parameters, sort_keys=True, default=str)}

Leakage prevention
Features use only information available on or before date t. Rolling windows are trailing, never centered. Future closes, targets, next-day returns and test-period labels are excluded from model.fit(). Missing rolling values are dropped, not future-filled. Test results are not used for tuning.

Evaluation
Held-out Accuracy, Precision, Recall, F1 and ROC-AUC (when both actual classes are present). UP/1 is the positive class.

Naive baseline
Today's positive daily return predicts tomorrow UP; otherwise it predicts DOWN. It is evaluated on exactly the same test observations.

Trading strategy
Prediction UP means invested long for Close(t) to Close(t+1); DOWN means cash. Strategy_Return = Prediction * Next_Day_Return. Buy-and-hold uses Next_Day_Return. Transaction cost per position change: {transaction_cost:.4f}%. A position change is an entry or exit. Metrics assume 252 trading days and a 0% risk-free rate. Starting portfolio value is 1.0.

Limitations
This is a simplified academic experiment. Feature importance is not causation. Results are sensitive to the instrument, date range and market regime. The backtest omits slippage, taxes, liquidity constraints and intraday execution effects. This output is not financial advice.
"""


def _write_df(ws, frame: pd.DataFrame, start_row: int = 1) -> None:
    for j, column in enumerate(frame.columns, 1):
        cell = ws.cell(start_row, j, str(column))
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(frame.itertuples(index=False, name=None), start_row + 1):
        for j, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif isinstance(value, np.generic):
                value = value.item()
            ws.cell(i, j, value)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(frame.columns))}{start_row + len(frame)}"


def _style_sheet(ws, freeze: str | None = "A2") -> None:
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for column in ws.columns:
        values = [str(c.value) if c.value is not None else "" for c in column[:200]]
        width = min(max(max((len(x) for x in values), default=8) + 2, 11), 34)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def _title(ws, text: str) -> None:
    ws["A1"] = text
    ws["A1"].font = Font(size=16, bold=True, color=NAVY)


def _section(ws, row: int, text: str) -> None:
    ws.cell(row, 1, text)
    ws.cell(row, 1).font = Font(bold=True, color=WHITE)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)


def create_excel_report(prepared: PreparedData, results: dict, parameters: dict, transaction_cost: float) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Random Forest Stock Direction Prediction — Research Summary")
    summary = [
        ("Dataset date range", f"{prepared.raw_data.Date.min():%Y-%m-%d} to {prepared.raw_data.Date.max():%Y-%m-%d}"),
        ("Training period", prepared.training_period), ("Test period", prepared.test_period),
        ("Raw observations", prepared.raw_rows), ("Usable observations", prepared.usable_rows),
        ("Training observations", len(prepared.train_data)), ("Test observations", len(prepared.test_evaluation)),
        ("Selected features", ", ".join(FEATURE_LABELS[f] for f in prepared.selected_features)),
        ("Model type", "RandomForestClassifier"), ("Random Forest parameters", json.dumps(parameters, default=str)),
        ("Transaction cost per position change", transaction_cost / 100),
        ("Generation timestamp", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")),
    ]
    for row, (label, value) in enumerate(summary, 3):
        ws.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        ws.cell(row, 2, value)
    ws["B13"].number_format = "0.0000%"
    row = 16
    _section(ws, row, "Classification Performance")
    for i, (name, value) in enumerate(results["metrics"].items(), row + 1):
        ws.cell(i, 1, name); ws.cell(i, 2, value if value is not None else "N/A")
        if value is not None: ws.cell(i, 2).number_format = "0.00%" if name != "ROC-AUC" else "0.000"
    row = 23
    _section(ws, row, "Benchmark")
    rf_acc = results["metrics"]["Accuracy"]; base_acc = results["baseline_metrics"]["Accuracy"]
    for i, item in enumerate([("Random Forest Accuracy", rf_acc), ("Naive Baseline Accuracy", base_acc), ("Difference (percentage points)", (rf_acc-base_acc)*100)], row+1):
        ws.cell(i, 1, item[0]); ws.cell(i, 2, item[1]); ws.cell(i, 2).number_format = "0.00%" if i < row+3 else "0.00"
    row = 29
    _section(ws, row, "Trading Performance")
    ws.cell(row+1, 1, "Metric"); ws.cell(row+1, 2, "RF Strategy"); ws.cell(row+1, 3, "Buy & Hold")
    for cell in ws[row+1]: cell.font = Font(bold=True, color=NAVY)
    for i, metric in enumerate(results["trading_metrics"]["RF Strategy"], row+2):
        ws.cell(i, 1, metric); ws.cell(i, 2, results["trading_metrics"]["RF Strategy"][metric]); ws.cell(i, 3, results["trading_metrics"]["Buy & Hold"][metric])
        if metric != "Sharpe Ratio": ws.cell(i, 2).number_format = ws.cell(i, 3).number_format = "0.00%"
    _style_sheet(ws, "A3")

    for name, frame in [("Raw Data", prepared.raw_data), ("Engineered Dataset", prepared.engineered_data)]:
        sheet = wb.create_sheet(name); _write_df(sheet, frame); _style_sheet(sheet)

    sheet = wb.create_sheet("Test Predictions")
    prediction_export = results["predictions"][["Date", "Close", "Actual_Next_Close", "Actual_Target", "Actual_Direction", "Prediction", "Predicted_Direction", "Probability_DOWN", "Probability_UP", "Correct", "Next_Day_Return", "Strategy_Return", "Buy_Hold_Return"]]
    _write_df(sheet, prediction_export); _style_sheet(sheet)
    for col in [8, 9, 11, 12, 13]:
        for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2): cell[0].number_format = "0.00%"
    correct_col = 10
    sheet.conditional_formatting.add(f"J2:J{len(prediction_export)+1}", CellIsRule(operator="equal", formula=["TRUE"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    sheet.conditional_formatting.add(f"J2:J{len(prediction_export)+1}", CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor=PALE_RED)))

    sheet = wb.create_sheet("Evaluation Metrics"); _title(sheet, "Held-Out Classification Metrics")
    _write_df(sheet, metrics_frame(results), 3)
    cm = results["confusion_matrix"]
    sheet["A11"] = "Confusion Matrix"; sheet["A11"].font = Font(bold=True, color=NAVY)
    sheet.append([]); sheet.append(["Actual / Predicted", "Predicted DOWN", "Predicted UP"])
    sheet.append(["Actual DOWN", int(cm[0,0]), int(cm[0,1])]); sheet.append(["Actual UP", int(cm[1,0]), int(cm[1,1])]); _style_sheet(sheet, "A4")

    sheet = wb.create_sheet("Benchmark Comparison"); _write_df(sheet, results["benchmark_comparison"]); _style_sheet(sheet)
    sheet = wb.create_sheet("Feature Importance"); _write_df(sheet, results["feature_importance"][["Display Feature", "Importance", "Rank"]].rename(columns={"Display Feature":"Feature"})); _style_sheet(sheet)
    chart = BarChart(); chart.title = "Random Forest Feature Importance"; chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=len(results["feature_importance"])+1), titles_from_data=True); chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(results["feature_importance"])+1)); sheet.add_chart(chart, "E2")

    sheet = wb.create_sheet("Trading Performance"); _write_df(sheet, results["trading_performance"]); start = len(results["trading_performance"])+4
    sheet.cell(start,1,"Summary Metric"); sheet.cell(start,2,"RF Strategy"); sheet.cell(start,3,"Buy & Hold")
    for c in sheet[start]: c.fill=PatternFill("solid",fgColor=NAVY); c.font=Font(color=WHITE,bold=True)
    for i, metric in enumerate(results["trading_metrics"]["RF Strategy"], start+1):
        sheet.cell(i,1,metric); sheet.cell(i,2,results["trading_metrics"]["RF Strategy"][metric]); sheet.cell(i,3,results["trading_metrics"]["Buy & Hold"][metric])
    line = LineChart(); line.title = "RF Strategy vs Buy-and-Hold"; line.y_axis.title="Portfolio Value"; line.x_axis.title="Date"
    line.add_data(Reference(sheet,min_col=4,max_col=5,min_row=1,max_row=len(results["trading_performance"])+1),titles_from_data=True); line.set_categories(Reference(sheet,min_col=1,min_row=2,max_row=len(results["trading_performance"])+1)); sheet.add_chart(line,"I2"); _style_sheet(sheet)

    sheet = wb.create_sheet("Methodology"); _title(sheet, "Methodology and Assumptions")
    for i, paragraph in enumerate(methodology_text(prepared, parameters, transaction_cost).splitlines(), 3):
        sheet.cell(i,1,paragraph); sheet.cell(i,1).alignment=Alignment(wrap_text=True,vertical="top")
        if paragraph and not paragraph.endswith(".") and len(paragraph)<40: sheet.cell(i,1).font=Font(bold=True,color=NAVY)
    sheet.column_dimensions["A"].width=110; sheet.sheet_view.showGridLines=False

    for sheet in wb.worksheets:
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if cell.row > 1 and cell.value is not None and cell.border == Border():
                    cell.border = Border(bottom=THIN_GRAY)
    output = BytesIO(); wb.save(output); return output.getvalue()


def _chart_png(kind: str, results: dict) -> bytes:
    fig, ax = plt.subplots(figsize=(7, 4), dpi=150)
    if kind == "confusion_matrix":
        matrix = results["confusion_matrix"]; ax.imshow(matrix, cmap="Blues")
        for (i,j), value in np.ndenumerate(matrix): ax.text(j,i,str(value),ha="center",va="center")
        ax.set_xticks([0,1],["Predicted DOWN","Predicted UP"]); ax.set_yticks([0,1],["Actual DOWN","Actual UP"]); ax.set_title("Confusion Matrix")
    elif kind == "feature_importance":
        data=results["feature_importance"].sort_values("Importance"); ax.barh(data["Display Feature"],data["Importance"],color="#3478A8"); ax.set_title("Feature Importance")
    else:
        data=results["trading_performance"]; ax.plot(data["Date"],data["RF Cumulative Value"],label="RF Strategy"); ax.plot(data["Date"],data["Buy-and-Hold Cumulative Value"],label="Buy & Hold"); ax.legend(); ax.set_title("RF Strategy vs Buy-and-Hold"); ax.set_ylabel("Portfolio Value")
    fig.tight_layout(); out=BytesIO(); fig.savefig(out,format="png",bbox_inches="tight"); plt.close(fig); return out.getvalue()


def create_research_package(prepared: PreparedData, results: dict, model, parameters: dict, transaction_cost: float) -> bytes:
    excel = create_excel_report(prepared, results, parameters, transaction_cost)
    configuration = _configuration(prepared, parameters)
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        prefix = "research_package/"
        files = {
            "engineered_dataset.csv": dataframe_csv(prepared.engineered_data),
            "test_predictions.csv": dataframe_csv(results["predictions"]),
            "classification_metrics.csv": dataframe_csv(metrics_frame(results)),
            "benchmark_comparison.csv": dataframe_csv(results["benchmark_comparison"]),
            "feature_importance.csv": dataframe_csv(results["feature_importance"]),
            "trading_performance.csv": dataframe_csv(results["trading_performance"]),
            "model_configuration.json": json.dumps(configuration, indent=2, default=str).encode(),
            "methodology.txt": methodology_text(prepared, parameters, transaction_cost).encode(),
            "random_forest_model.joblib": model_bytes(model),
            "random_forest_research_report.xlsx": excel,
            "charts/confusion_matrix.png": _chart_png("confusion_matrix", results),
            "charts/feature_importance.png": _chart_png("feature_importance", results),
            "charts/strategy_vs_buy_hold.png": _chart_png("strategy", results),
        }
        for name, content in files.items(): archive.writestr(prefix + name, content)
    return output.getvalue()
